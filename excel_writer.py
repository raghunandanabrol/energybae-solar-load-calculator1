import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import io
from datetime import datetime


# ── Styles ────────────────────────────────────────────────────────────
GREEN_DARK  = "1B5E20"
GREEN_MID   = "388E3C"
GREEN_LIGHT = "C8E6C9"
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F5F5F5"
YELLOW      = "FFF9C4"


def _header_font(size=12, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _normal_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_header_row(ws, row, cols, text, bg=GREEN_DARK):
    start_col, end_col = cols
    ws.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row,   end_column=end_col
    )
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font      = _header_font(13, True, WHITE)
    cell.fill      = _fill(bg)
    cell.alignment = _center()
    cell.border    = _border()


def _write_row(ws, row, label, value, unit="",
               label_col=1, value_col=2, unit_col=3, highlight=False):
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font      = _normal_font(10, True)
    lc.fill      = _fill(GREEN_LIGHT) if highlight else _fill(GRAY_LIGHT)
    lc.alignment = _left()
    lc.border    = _border()

    vc = ws.cell(row=row, column=value_col, value=value)
    vc.font      = _normal_font(10, False)
    vc.fill      = _fill(YELLOW) if highlight else _fill(WHITE)
    vc.alignment = _center()
    vc.border    = _border()

    uc = ws.cell(row=row, column=unit_col, value=unit)
    uc.font      = _normal_font(9, False, "757575")
    uc.fill      = _fill(WHITE)
    uc.alignment = _center()
    uc.border    = _border()


# ── Sheet 1: Customer Info ─────────────────────────────────────────────
def _build_customer_sheet(wb, extracted: dict):
    ws = wb.active
    ws.title = "Customer Info"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15

    ws.row_dimensions[1].height = 40
    _write_header_row(ws, 1, (1, 3),
        "ENERGYBAE — Solar Load Calculator | Customer Bill Data")

    ws.row_dimensions[2].height = 25
    _write_header_row(ws, 2, (1, 3),
        f"Generated: {datetime.now().strftime('%d %B %Y  %I:%M %p')}",
        bg=GREEN_MID)

    _write_header_row(ws, 3, (1, 3), "CUSTOMER & BILL DETAILS", bg=GREEN_MID)

    fields = [
        ("Consumer Name",        extracted.get("consumer_name"),         ""),
        ("Consumer Number",      extracted.get("consumer_number"),        ""),
        ("Meter Number",         extracted.get("meter_number"),           ""),
        ("Billing Address",      extracted.get("billing_address"),        ""),
        ("Distribution Company", extracted.get("distribution_company"),   ""),
        ("Billing Month",        extracted.get("billing_month"),          ""),
        ("Billing Period From",  extracted.get("billing_period_from"),    ""),
        ("Billing Period To",    extracted.get("billing_period_to"),      ""),
        ("Number of Days",       extracted.get("number_of_days"),         "days"),
        ("Tariff Category",      extracted.get("tariff_category"),        ""),
        ("Phase Type",           extracted.get("phase_type"),             ""),
        ("Sanctioned Load",      extracted.get("sanctioned_load_kw"),     "kW"),
        ("Contract Demand",      extracted.get("contract_demand_kva"),    "kVA"),
        ("Previous Reading",     extracted.get("previous_reading"),       "units"),
        ("Current Reading",      extracted.get("current_reading"),        "units"),
        ("Units Consumed",       extracted.get("units_consumed_kwh"),     "kWh"),
        ("Monthly Avg Units",    extracted.get("monthly_avg_units"),      "kWh"),
        ("Energy Charges",       extracted.get("energy_charges"),         "Rs"),
        ("Fixed Charges",        extracted.get("fixed_charges"),          "Rs"),
        ("Tax Amount",           extracted.get("tax_amount"),             "Rs"),
        ("Total Bill Amount",    extracted.get("total_bill_amount"),      "Rs"),
        ("Power Factor",         extracted.get("power_factor"),           ""),
    ]

    highlight_labels = ["Units Consumed", "Total Bill Amount", "Sanctioned Load"]
    for i, (label, value, unit) in enumerate(fields):
        r = 4 + i
        ws.row_dimensions[r].height = 20
        _write_row(ws, r, label, value, unit,
                   highlight=(label in highlight_labels))


# ── Sheet 2: Solar Calculation ─────────────────────────────────────────
def _build_solar_sheet(wb, extracted: dict, calcs: dict):
    ws = wb.create_sheet("Solar Calculation")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18

    row = 1
    _write_header_row(ws, row, (1, 3),
        "SOLAR SYSTEM SIZING — ENERGYBAE CALCULATOR")

    row += 1
    _write_header_row(ws, row, (1, 3), "A. CONSUMPTION ANALYSIS", bg=GREEN_MID)

    for label, value, unit in [
        ("Monthly Consumption",       calcs["monthly_consumption_kwh"], "kWh/month"),
        ("Daily Average Consumption", calcs["daily_consumption_kwh"],   "kWh/day"),
        ("Per Unit Rate (estimated)", calcs["per_unit_rate_inr"],       "Rs/kWh"),
        ("Current Monthly Bill",      calcs["monthly_bill_inr"],        "Rs"),
    ]:
        row += 1
        ws.row_dimensions[row].height = 20
        _write_row(ws, row, label, value, unit)

    row += 1
    _write_header_row(ws, row, (1, 3), "B. SOLAR SYSTEM SIZING", bg=GREEN_MID)

    highlight_solar = ["Solar Capacity Recommended", "Number of Panels (400Wp each)"]
    for label, value, unit in [
        ("Peak Sun Hours (India avg)",    calcs["peak_sun_hours"],        "hrs/day"),
        ("System Efficiency",             calcs["system_efficiency_pct"], "%"),
        ("Solar Capacity Required",       calcs["solar_kwp_required"],    "kWp"),
        ("Solar Capacity Recommended",    calcs["solar_kwp_recommended"], "kWp"),
        ("Number of Panels (400Wp each)", calcs["num_panels_400wp"],      "panels"),
        ("Inverter Size",                 calcs["inverter_size_kva"],     "kVA"),
        ("Annual Energy Generation",      calcs["annual_generation_kwh"], "kWh/year"),
        ("Solar Coverage",                calcs["solar_coverage_pct"],    "%"),
    ]:
        row += 1
        ws.row_dimensions[row].height = 20
        _write_row(ws, row, label, value, unit,
                   highlight=(label in highlight_solar))

    row += 1
    _write_header_row(ws, row, (1, 3), "C. ENVIRONMENTAL IMPACT", bg="2E7D32")

    for label, value, unit in [
        ("CO2 Saved Annually",       calcs["co2_saved_kg_annual"], "kg/year"),
        ("Equivalent Trees Planted", calcs["trees_equivalent"],    "trees"),
    ]:
        row += 1
        ws.row_dimensions[row].height = 20
        _write_row(ws, row, label, value, unit)


# ── Sheet 3: Financial Summary ─────────────────────────────────────────
def _build_financial_sheet(wb, calcs: dict):
    ws = wb.create_sheet("Financial Summary")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18

    row = 1
    _write_header_row(ws, row, (1, 3),
        "FINANCIAL ANALYSIS — ROI & SAVINGS")

    row += 1
    _write_header_row(ws, row, (1, 3), "A. SYSTEM COST & SUBSIDY", bg=GREEN_MID)

    for label, value, unit in [
        ("Total System Cost (before subsidy)", calcs["total_system_cost_inr"],  "Rs"),
        ("Government Subsidy (PM Surya Ghar)", calcs["govt_subsidy_inr"],       "Rs"),
        ("Net Cost After Subsidy",             calcs["net_cost_after_subsidy"], "Rs"),
    ]:
        row += 1
        ws.row_dimensions[row].height = 22
        _write_row(ws, row, label, value, unit,
                   highlight=(label == "Net Cost After Subsidy"))

    row += 1
    _write_header_row(ws, row, (1, 3), "B. SAVINGS & RETURNS", bg=GREEN_MID)

    for label, value, unit in [
        ("Annual Electricity Savings",     calcs["annual_savings_inr"],                  "Rs/year"),
        ("Monthly Savings (estimated)",    round(calcs["annual_savings_inr"] / 12, 0),   "Rs/month"),
        ("Payback Period",                 calcs["payback_period_years"],                "years"),
        ("25-Year ROI",                    calcs["roi_25_years_pct"],                    "%"),
        ("Lifetime Savings (25 years)",    round(calcs["annual_savings_inr"] * 25, 0),   "Rs"),
    ]:
        row += 1
        ws.row_dimensions[row].height = 22
        _write_row(ws, row, label, value, unit,
                   highlight=(label in ["Payback Period", "25-Year ROI"]))

    # ── Year-wise Table ───────────────────────────────────────────────
    row += 2
    _write_header_row(ws, row, (1, 3),
        "C. YEAR-WISE SAVINGS PROJECTION (25 YEARS)", bg=GREEN_MID)

    row += 1
    for col, hdr in enumerate(["Year", "Annual Saving (Rs)", "Cumulative Saving (Rs)"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font      = _header_font(10, True, WHITE)
        c.fill      = _fill("558B2F")
        c.alignment = _center()
        c.border    = _border()

    annual     = calcs["annual_savings_inr"]
    cumulative = 0
    table_start_row = row + 1

    for yr in range(1, 26):
        row += 1
        yr_saving  = round(annual * (1.03 ** (yr - 1)), 0)
        cumulative += yr_saving
        ws.row_dimensions[row].height = 18
        for col, val in enumerate([yr, yr_saving, round(cumulative, 0)], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = _normal_font(9)
            c.fill      = _fill(GRAY_LIGHT) if yr % 2 == 0 else _fill(WHITE)
            c.alignment = _center()
            c.border    = _border()

    # ── Bar Chart ─────────────────────────────────────────────────────
    chart = BarChart()
    chart.type          = "col"
    chart.title         = "25-Year Cumulative Savings"
    chart.y_axis.title  = "Cumulative Savings (Rs)"
    chart.x_axis.title  = "Year"
    chart.width         = 20
    chart.height        = 12

    data_ref = Reference(ws, min_col=3, max_col=3,
                         min_row=table_start_row, max_row=row)
    cats_ref = Reference(ws, min_col=1, max_col=1,
                         min_row=table_start_row, max_row=row)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "4CAF50"
    ws.add_chart(chart, f"A{row + 3}")


# ── Main Function ─────────────────────────────────────────────────────
def generate_excel_report(extracted: dict, calcs: dict) -> bytes:
    wb = openpyxl.Workbook()
    _build_customer_sheet(wb, extracted)
    _build_solar_sheet(wb, extracted, calcs)
    _build_financial_sheet(wb, calcs)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()